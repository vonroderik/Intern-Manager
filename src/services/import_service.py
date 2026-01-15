import csv
import openpyxl
from pathlib import Path
from services.intern_service import InternService
from services.venue_service import VenueService
from services.document_service import DocumentService
from core.models.venue import Venue
from core.models.intern import Intern

class ImportService:
    def __init__(
        self,
        intern_service: InternService,
        venue_service: VenueService,
        document_service: DocumentService,
    ):
        self.intern_service = intern_service
        self.venue_service = venue_service
        self.document_service = document_service

    def read_file(self, filename: str | Path) -> None:
        """
        Lê um arquivo (CSV ou Excel) e importa os dados.
        Detecta automaticamente o formato pela extensão.
        """
        path = Path(filename)
        suffix = path.suffix.lower()

        rows = []

        try:
            if suffix == ".csv":
                rows = self._read_csv(path)
            elif suffix in [".xlsx", ".xls"]:
                rows = self._read_excel(path)
            else:
                raise ValueError("Formato não suportado. Use .csv ou .xlsx")

            self._process_data(rows)

        except Exception as e:
            print(f"ERRO NA IMPORTAÇÃO: {e}")
            raise e

    def _read_csv(self, path: Path) -> list[dict]:
        """Lê CSV tentando diferentes encodings."""
        rows = []
        encodings = ["utf-8-sig", "latin-1", "cp1252"]
        
        for enc in encodings:
            try:
                with open(path, "r", newline="", encoding=enc) as f:
                    sample = f.read(2048)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample)
                        delimiter = dialect.delimiter
                    except csv.Error:
                        delimiter = ";" 

                    reader = csv.DictReader(f, delimiter=delimiter)
                    rows = list(reader)
                return rows
            except UnicodeDecodeError:
                continue
            except Exception as e:
                raise e
        
        raise ValueError("Não foi possível decodificar o arquivo CSV.")

    def _read_excel(self, path: Path) -> list[dict]:
        """Lê Excel e usa a primeira linha como cabeçalho."""
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active
        
        if sheet is None:
            raise ValueError("O arquivo Excel não possui uma planilha ativa.")
        
        rows = []
        headers = []
        
        # iter_rows retorna células. values_only=True retorna os valores direto.
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                # Cabeçalho
                headers = [str(cell).strip() if cell else f"col_{j}" for j, cell in enumerate(row)]
                continue
            
            row_dict = {}
            has_data = False
            for j, value in enumerate(row):
                if j < len(headers):
                    val_str = str(value).strip() if value is not None else ""
                    row_dict[headers[j]] = val_str
                    if val_str: has_data = True
            
            if has_data:
                rows.append(row_dict)
                
        return rows

    def _process_data(self, rows: list[dict]):
        processed_venues = set()
        processed_interns = set()
        venue_id_map: dict[str, int] = {}
        line_count = 0

        for row in rows:
            line_count += 1
            # Normaliza chaves para minúsculo para evitar erro de digitação no header
            safe_row = {k.lower().strip(): v for k, v in row.items()}

            venue_name = safe_row.get("local", "").strip()
            intern_name = safe_row.get("nome", "").strip()
            ra_raw = safe_row.get("ra", "").strip()

            if not intern_name or not ra_raw:
                continue

            # --- Processa Local ---
            current_venue_id: int | None = None
            raw_sup_email = safe_row.get("email_supervisor")
            email_sup = raw_sup_email.strip() if raw_sup_email else None

            # Tenta criar ou buscar local
            if venue_name and venue_name not in processed_venues:
                existing_venue = self.venue_service.repo.get_by_name(venue_name)
                venue_data = {
                    "venue_name": venue_name,
                    "supervisor_name": safe_row.get("nome_supervisor", "").strip(),
                    "supervisor_email": email_sup,
                    "supervisor_phone": safe_row.get("telefone_supervisor", "").strip(),
                }

                if existing_venue:
                    self.venue_service.update_venue(Venue(venue_id=existing_venue.venue_id, **venue_data))
                    current_venue_id = existing_venue.venue_id
                else:
                    current_venue_id = self.venue_service.add_new_venue(Venue(**venue_data))
                
                # Fallback
                if current_venue_id is None:
                    v = self.venue_service.repo.get_by_name(venue_name)
                    if v: current_venue_id = v.venue_id

                if current_venue_id is not None:
                    processed_venues.add(venue_name)
                    venue_id_map[venue_name] = current_venue_id
            
            elif venue_name:
                current_venue_id = venue_id_map.get(venue_name)
                if not current_venue_id:
                    v = self.venue_service.repo.get_by_name(venue_name)
                    if v: current_venue_id = v.venue_id

            # --- Processa Aluno ---
            if intern_name in processed_interns:
                continue

            existing_intern = self.intern_service.repo.get_by_name(intern_name)
            intern_data = {
                "name": intern_name,
                "registration_number": str(ra_raw),
                "venue_id": current_venue_id,
                "term": safe_row.get("periodo", "").strip(),
                "email": safe_row.get("email", None),
                "start_date": safe_row.get("data_inicio", "").strip(),
                "end_date": safe_row.get("data_fim", "").strip(),
                "working_hours": safe_row.get("horarios", "").strip(),
            }

            if existing_intern:
                self.intern_service.update_intern(Intern(intern_id=existing_intern.intern_id, **intern_data))
            else:
                new_intern_id = self.intern_service.add_new_intern(Intern(**intern_data))
                # Cria documentos iniciais
                if new_intern_id:
                    try:
                        self.document_service.create_initial_documents_batch(new_intern_id)
                    except Exception:
                        pass

            processed_interns.add(intern_name)