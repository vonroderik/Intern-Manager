import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill

class ExportService:
    def __init__(self, db_connector):
        self.db = db_connector

    def export_to_excel(self, filepath: str):
        """
        Exporta todas as tabelas principais para um arquivo Excel (.xlsx).
        """
        tables = [
            "interns", "venues", "documents", 
            "observations", "meetings", "grades", 
            "evaluation_criteria"
        ]

        wb = openpyxl.Workbook()
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        # CORREÇÃO: Acessa o atributo .conn diretamente da sua classe DatabaseConnector
        conn = self.db.conn
        if not conn:
            raise RuntimeError("Banco de dados não conectado.")
            
        cursor = conn.cursor()

        try:
            for table_name in tables:
                self._export_table(wb, cursor, table_name)
            
            wb.save(filepath)
            print(f"Exportação concluída: {filepath}")

        except Exception as e:
            print(f"Erro na exportação: {e}")
            raise e
        finally:
            cursor.close()

    def _export_table(self, wb, cursor, table_name):
        try:
            cursor.execute(f"SELECT * FROM {table_name}")
            rows = cursor.fetchall()
            
            # Pega nomes das colunas
            columns = [description[0] for description in cursor.description]
            
            ws = wb.create_sheet(title=table_name.capitalize())
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

            ws.append(columns)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            for row in rows:
                ws.append(list(row))

            # Ajuste de largura
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
                
        except sqlite3.OperationalError:
            print(f"Aviso: Tabela '{table_name}' não encontrada.")